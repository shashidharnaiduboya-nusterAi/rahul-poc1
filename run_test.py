"""
run_test.py -- Standalone E2E Test Runner with Groq LLM (v2)
=============================================================
Pipeline:
  1. Parse XLSX ground truth (alert → PG doc → section)
  2. Index all B&F PG docs (embedding + BM25)
  3. Hybrid retrieval: semantic + BM25 combined scoring
  4. Groq/Llama guardrail: LLM filters out false positives
  5. Section-level matching: Groq identifies impacted sections
  6. Evaluation: doc-level + section-level P / R / F1

Usage:
    python3 run_test.py
    python3 run_test.py --top-k 30 --threshold 0.15 --max-alerts 10
"""

from __future__ import annotations

import argparse, json, math, os, re, sys, time, uuid
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
from dotenv import load_dotenv
from groq import Groq
from rank_bm25 import BM25Okapi

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
XLSX_PATH = BASE_DIR / "ACP PG Golden Data Set.xlsx"
PG_DOCS_DIR = BASE_DIR / "ACP-PG_3PAs_OV-PN-CL-SF" / "BANKINGANDFINANCE"
TEST_QDRANT_DIR = BASE_DIR / "data" / "test_qdrant"
REPORT_DIR = BASE_DIR / "data" / "test_reports"

GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
GROQ_MODEL = "llama-3.3-70b-versatile"
COLLECTION_NAME = "test_pg_docs"

_STOPWORDS = frozenset(
    "the a an and or but in on at to for of with by from as is was are were be "
    "been being have has had do does did will would could should may might shall "
    "that this these those it its their they he she we i you not no".split()
)


# ═══════════════════════════════════════════════════════════════════
# Data classes
# ═══════════════════════════════════════════════════════════════════
@dataclass
class GroundTruthEntry:
    alert_lni: str; alert_title: str; key_terms: str; pg_title: str
    echo_id: str; pg_type: str; section_heading: str
    complexity: str; rationale: str

@dataclass
class PGDoc:
    doc_id: str; echo_id: str; title: str; practice_area: str
    jurisdiction: str; sections: list[dict] = field(default_factory=list)
    full_text: str = ""; source_file: str = ""

@dataclass
class AlertResult:
    alert_lni: str; alert_title: str; key_terms: str
    expected_doc_ids: set; expected_sections: dict
    predicted_doc_ids: set = field(default_factory=set)
    predicted_sections: dict = field(default_factory=dict)
    retrieval_scores: dict = field(default_factory=dict)
    guardrail_kept: set = field(default_factory=set)
    doc_tp: int = 0; doc_fp: int = 0; doc_fn: int = 0
    sec_tp: int = 0; sec_fp: int = 0; sec_fn: int = 0


# ═══════════════════════════════════════════════════════════════════
# 1. XLSX Ground Truth
# ═══════════════════════════════════════════════════════════════════
def parse_ground_truth(xlsx_path: Path):
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    entries = []
    ws = wb["Banking and Finance Alerts"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row) + [None] * 22
        if not vals[3]: continue
        echo_raw = vals[7]
        eid = str(int(echo_raw)) if isinstance(echo_raw, (int, float)) else str(echo_raw or "")
        entries.append(GroundTruthEntry(
            alert_lni=str(vals[3]).strip(), alert_title=str(vals[2] or ""),
            key_terms=str(vals[5] or ""), pg_title=str(vals[6] or ""),
            echo_id=eid, pg_type=str(vals[10] or ""),
            section_heading=str(vals[12] or ""), complexity=str(vals[13] or ""),
            rationale=str(vals[14] or "")))
    no_impact = []
    if "B and F Alerts No impact" in wb.sheetnames:
        ws2 = wb["B and F Alerts No impact"]
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
            vals = list(row) + [None] * 5
            if vals[2]: no_impact.append(str(vals[2]).strip())
    wb.close()
    return entries, no_impact

def build_alert_ground_truth(entries):
    alerts = {}
    for e in entries:
        if e.alert_lni not in alerts:
            alerts[e.alert_lni] = {"title": e.alert_title, "key_terms": e.key_terms,
                                   "expected_docs": {}, "expected_sections": {}}
        a = alerts[e.alert_lni]
        if e.echo_id:
            a["expected_docs"][e.echo_id] = e.pg_title
            a["expected_sections"].setdefault(e.echo_id, [])
            if e.section_heading:
                a["expected_sections"][e.echo_id].append(e.section_heading)
    return alerts


# ═══════════════════════════════════════════════════════════════════
# 2. PG Doc XML Parser
# ═══════════════════════════════════════════════════════════════════
def _et(elem):
    parts = []
    if elem.text and elem.text.strip(): parts.append(elem.text.strip())
    for ch in elem:
        t = _et(ch)
        if t: parts.append(t)
        if ch.tail and ch.tail.strip(): parts.append(ch.tail.strip())
    return " ".join(parts)

def _ln(tag): return tag.split("}", 1)[1] if "}" in tag else tag

def parse_pg_xml(xml_path: Path) -> Optional[PGDoc]:
    import xml.etree.ElementTree as ET
    try: root = ET.parse(str(xml_path)).getroot()
    except Exception: return None
    echo_id = xml_path.stem.split("_")[-1]
    doc_id = echo_id
    title = ""
    for el in root.iter():
        if _ln(el.tag) == "front":
            for ch in el:
                if _ln(ch.tag) == "title":
                    for s in ch:
                        if _ln(s.tag) == "text" and s.text: title = s.text.strip(); break
            break
    if not title:
        for el in root.iter():
            if _ln(el.tag) == "document-title" and el.text: title = el.text.strip(); break
    sections, si = [], 0
    for el in root.iter():
        if _ln(el.tag) in ("clause", "inclusion", "section"):
            heading = ""
            for ch in el:
                if _ln(ch.tag) in ("heading", "title", "text"):
                    h = _et(ch).strip()
                    if h and len(h) < 300: heading = h; break
            text = _et(el).strip()
            if text and len(text) > 30:
                sections.append({"id": f"sec_{si}", "heading": heading, "text": text[:3000]}); si += 1
    if not sections:
        for el in root.iter():
            if _ln(el.tag) in ("para", "pgrp", "p"):
                text = _et(el).strip()
                if text and len(text) > 30:
                    sections.append({"id": f"para_{si}", "heading": "", "text": text[:3000]}); si += 1
    headings = [s["heading"] for s in sections if s.get("heading")]
    heading_block = " | ".join(headings[:50])
    lead_paras = [s["text"][:300] for s in sections[:10]]
    full_text = f"{title}\n{title}\n\nSections: {heading_block}\n\n" + "\n\n".join(lead_paras)
    full_text = full_text[:5000]
    return PGDoc(doc_id=doc_id, echo_id=echo_id, title=title, practice_area="",
                 jurisdiction="", sections=sections, full_text=full_text,
                 source_file=str(xml_path))


# ═══════════════════════════════════════════════════════════════════
# 3. BM25 Index
# ═══════════════════════════════════════════════════════════════════
def _tokenize(text: str) -> list[str]:
    return [t for t in re.findall(r"[a-zA-Z]{2,}", text.lower()) if t not in _STOPWORDS]

class BM25Index:
    def __init__(self, docs: list[PGDoc]):
        self.echo_ids = [d.echo_id for d in docs]
        corpus = [_tokenize(d.full_text) for d in docs]
        self.bm25 = BM25Okapi(corpus)
        self._echo_to_idx = {eid: i for i, eid in enumerate(self.echo_ids)}

    def query(self, text: str, top_k: int = 50) -> dict[str, float]:
        tokens = _tokenize(text)
        if not tokens:
            return {}
        scores = self.bm25.get_scores(tokens)
        max_s = scores.max() if scores.max() > 0 else 1.0
        top_idx = np.argsort(scores)[::-1][:top_k]
        return {self.echo_ids[i]: float(scores[i] / max_s) for i in top_idx if scores[i] > 0}


# ═══════════════════════════════════════════════════════════════════
# 4. Embedding + Qdrant Indexing
# ═══════════════════════════════════════════════════════════════════
def index_pg_docs(docs, qdrant_path, batch_size=64):
    from sentence_transformers import SentenceTransformer
    from qdrant_client import QdrantClient
    from qdrant_client.models import Distance, VectorParams, PointStruct

    print("[Index] Loading all-MiniLM-L6-v2 (384-dim)...")
    model = SentenceTransformer("all-MiniLM-L6-v2")
    qdrant_path.mkdir(parents=True, exist_ok=True)
    qc = QdrantClient(path=str(qdrant_path))
    if COLLECTION_NAME in [c.name for c in qc.get_collections().collections]:
        qc.delete_collection(COLLECTION_NAME)
    qc.create_collection(COLLECTION_NAME,
                         vectors_config=VectorParams(size=384, distance=Distance.COSINE))
    texts = [d.full_text for d in docs]
    total = len(texts)
    print(f"[Index] Embedding {total} docs...", flush=True)
    all_vecs = []
    for s in range(0, total, batch_size):
        e = min(s + batch_size, total)
        all_vecs.extend(model.encode(texts[s:e], show_progress_bar=False, batch_size=batch_size))
        print(f"  {e}/{total}", flush=True)
    points = [PointStruct(id=str(uuid.uuid4()), vector=v.tolist(),
              payload={"doc_id": d.doc_id, "echo_id": d.echo_id, "title": d.title,
                       "source_file": d.source_file, "section_count": len(d.sections)})
              for d, v in zip(docs, all_vecs)]
    for s in range(0, len(points), 100):
        qc.upsert(COLLECTION_NAME, points[s:s+100])
    print(f"[Index] {qc.get_collection(COLLECTION_NAME).points_count} points indexed", flush=True)
    qc.close()
    return model


# ═══════════════════════════════════════════════════════════════════
# 5. Multi-Query Hybrid Retrieval + Keyword Reranker + Cross-Encoder
# ═══════════════════════════════════════════════════════════════════
def _semantic_query(model, qdrant_path, query_text, limit):
    from qdrant_client import QdrantClient
    vec = model.encode(query_text).tolist()
    qc = QdrantClient(path=str(qdrant_path))
    res = qc.query_points(COLLECTION_NAME, query=vec, limit=limit)
    qc.close()
    scores, payloads = {}, {}
    for pt in res.points:
        eid = pt.payload.get("echo_id", "")
        scores[eid] = pt.score
        payloads[eid] = pt.payload
    return scores, payloads


def _keyword_overlap_score(alert_terms: str, doc_title: str, doc_text: str) -> float:
    """Fraction of alert key-term tokens found in doc title/text."""
    terms = set(_tokenize(alert_terms))
    if not terms:
        return 0.0
    title_tokens = set(_tokenize(doc_title))
    text_tokens = set(_tokenize(doc_text[:2000]))
    title_hits = terms & title_tokens
    text_hits = terms & text_tokens
    title_frac = len(title_hits) / len(terms) if terms else 0
    text_frac = len(text_hits) / len(terms) if terms else 0
    return 0.6 * title_frac + 0.4 * text_frac


def hybrid_retrieve(model, qdrant_path, bm25_idx: BM25Index,
                    alert_title, key_terms, top_k=30, threshold=0.15,
                    alpha=0.6, echo_to_doc=None):
    """Multi-query hybrid retrieval with keyword reranking."""
    fetch_limit = top_k * 3

    q_combined = f"{alert_title} {key_terms}"
    q_title = alert_title
    q_terms = key_terms

    sem1, pl1 = _semantic_query(model, qdrant_path, q_combined, fetch_limit)
    sem2, pl2 = _semantic_query(model, qdrant_path, q_title, fetch_limit)
    sem3, pl3 = _semantic_query(model, qdrant_path, q_terms, fetch_limit)

    all_eids = set(sem1) | set(sem2) | set(sem3)
    payload_cache = {**pl3, **pl2, **pl1}

    sem_max = {}
    for eid in all_eids:
        sem_max[eid] = max(sem1.get(eid, 0), sem2.get(eid, 0), sem3.get(eid, 0))

    bm25_scores = bm25_idx.query(q_combined, top_k=fetch_limit)
    all_eids |= set(bm25_scores)

    combined = {}
    for eid in all_eids:
        ss = sem_max.get(eid, 0.0)
        bs = bm25_scores.get(eid, 0.0)
        base = alpha * ss + (1 - alpha) * bs

        doc = echo_to_doc.get(eid) if echo_to_doc else None
        doc_title = payload_cache.get(eid, {}).get("title", "")
        doc_text = doc.full_text if doc else doc_title
        kw_score = _keyword_overlap_score(key_terms, doc_title, doc_text)
        combined[eid] = base + 0.15 * kw_score

    ranked = sorted(combined.items(), key=lambda x: -x[1])[:top_k]
    hits = []
    for eid, score in ranked:
        if score < threshold:
            continue
        pl = payload_cache.get(eid, {})
        hits.append({
            "echo_id": eid, "doc_id": pl.get("doc_id", ""),
            "title": pl.get("title", ""), "score": round(score, 4),
            "sem_score": round(sem_max.get(eid, 0), 4),
            "bm25_score": round(bm25_scores.get(eid, 0), 4),
            "source_file": pl.get("source_file", ""),
        })
    return hits


_cross_encoder = None
def _get_cross_encoder():
    global _cross_encoder
    if _cross_encoder is None:
        from sentence_transformers import CrossEncoder
        print("  [CrossEncoder] Loading ms-marco-MiniLM-L-6-v2...", flush=True)
        _cross_encoder = CrossEncoder("cross-encoder/ms-marco-MiniLM-L-6-v2")
    return _cross_encoder


def cross_encoder_rerank(alert_text: str, hits: list[dict],
                         echo_to_doc: dict, top_k: int = 10) -> list[dict]:
    """Rerank retrieval hits using a cross-encoder for precise relevance."""
    if not hits:
        return []
    ce = _get_cross_encoder()
    pairs = []
    for h in hits:
        doc = echo_to_doc.get(h["echo_id"])
        doc_text = doc.full_text[:512] if doc else h.get("title", "")
        pairs.append((alert_text, doc_text))

    ce_scores = ce.predict(pairs)
    ce_min = float(min(ce_scores))
    ce_max = float(max(ce_scores))
    ce_range = ce_max - ce_min if ce_max > ce_min else 1.0

    for h, cs in zip(hits, ce_scores):
        ce_norm = (float(cs) - ce_min) / ce_range
        h["ce_score"] = round(float(cs), 4)
        h["ce_norm"] = round(ce_norm, 4)
        h["retrieval_score"] = h["score"]
        h["score"] = round(0.35 * h["retrieval_score"] + 0.65 * ce_norm, 4)

    hits.sort(key=lambda x: -x["score"])
    return hits[:top_k]


# ═══════════════════════════════════════════════════════════════════
# 6. Groq LLM Helpers (thread-safe via per-thread clients)
# ═══════════════════════════════════════════════════════════════════
import threading
_groq_local = threading.local()

def _get_groq() -> Groq:
    if not hasattr(_groq_local, "client"):
        _groq_local.client = Groq(api_key=GROQ_API_KEY)
    return _groq_local.client

def _groq_call(system: str, user: str, max_tokens: int = 250) -> str:
    client = _get_groq()
    for attempt in range(4):
        try:
            resp = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[{"role": "system", "content": system},
                          {"role": "user", "content": user}],
                temperature=0, max_completion_tokens=max_tokens)
            raw = resp.choices[0].message.content.strip()
            return re.sub(r"```(?:json)?", "", raw).strip().strip("`")
        except Exception as exc:
            if attempt < 3 and ("rate" in str(exc).lower() or "429" in str(exc)):
                wait = 3 * (2 ** attempt)
                time.sleep(wait)
            else:
                return ""
    return ""


# ═══════════════════════════════════════════════════════════════════
# 7. Guardrail (Groq LLM pre-filter)
# ═══════════════════════════════════════════════════════════════════
GUARDRAIL_SYS = """\
You are a senior Banking & Finance legal analyst performing a relevance pre-filter.
Given a case alert and a PG document title + summary, decide if this PG document
COULD PLAUSIBLY need updating due to the case.

Respond ONLY with JSON: {"relevant": true/false, "reason": "one sentence"}

Be GENEROUS at this stage — only reject if the PG doc is clearly unrelated to the
alert's legal topic. When in doubt, say true."""

def _guardrail_one(alert_title, key_terms, doc_title, doc_text_snippet):
    user = (f"ALERT: {alert_title}\nKEY TERMS: {key_terms}\n\n"
            f"PG DOCUMENT: {doc_title}\nSNIPPET: {doc_text_snippet[:800]}\n\n"
            f"Is this PG document plausibly relevant?")
    raw = _groq_call(GUARDRAIL_SYS, user, 120)
    try:
        r = json.loads(raw)
        return r.get("relevant", True)
    except Exception:
        return True  # pass through on error

def guardrail_filter(alert_title, key_terms, hits, echo_to_doc, workers=6):
    """Sequential guardrail with rate-limit pacing."""
    if not hits:
        return []
    kept = []
    for h in hits:
        doc = echo_to_doc.get(h["echo_id"])
        snippet = doc.full_text[:800] if doc else h.get("title", "")
        if _guardrail_one(alert_title, key_terms, h.get("title", ""), snippet):
            kept.append(h)
        time.sleep(0.5)
    kept.sort(key=lambda x: -x["score"])
    return kept


# ═══════════════════════════════════════════════════════════════════
# 8. Section Matching (Groq LLM)
# ═══════════════════════════════════════════════════════════════════
SEC_MATCH_SYS = """\
You are a senior Banking & Finance legal analyst. Given a case alert and a PG document section,
determine if this specific section needs updating due to the case.

Respond ONLY with JSON: {"is_impacted": true/false, "confidence": "HIGH"|"MEDIUM"|"LOW", "reason": "one sentence"}

STRICT: Only mark impacted if the case DIRECTLY affects this section's legal content.
DEFAULT to false unless clearly relevant."""

def _match_one_section(alert_title, key_terms, pg_title, sec):
    heading = sec.get("heading", "(no heading)")
    text = sec.get("text", "")[:1200]
    if len(text) < 30: return None
    user = (f"ALERT: {alert_title}\nKEY TERMS: {key_terms}\n\n"
            f"PG DOCUMENT: {pg_title}\nSECTION: {heading}\n"
            f"TEXT: {text}\n\nIs this section impacted?")
    raw = _groq_call(SEC_MATCH_SYS, user, 150)
    try:
        r = json.loads(raw)
        if r.get("is_impacted") and r.get("confidence") in ("HIGH", "MEDIUM"):
            return {"section_id": sec["id"], "heading": heading,
                    "confidence": r["confidence"], "reason": r.get("reason", "")}
    except Exception: pass
    return None

def match_sections_sequential(alert_title, key_terms, pg_title, sections):
    matched = []
    secs = [s for s in sections[:20] if len(s.get("text", "")) > 30]
    for s in secs:
        r = _match_one_section(alert_title, key_terms, pg_title, s)
        if r: matched.append(r)
        time.sleep(0.5)
    return matched


# ═══════════════════════════════════════════════════════════════════
# 9. Evaluation Helpers
# ═══════════════════════════════════════════════════════════════════
def _norm(s): return re.sub(r"[^a-z0-9 ]", "", s.lower()).strip()

def section_heading_matches(pred, exp):
    if not pred or not exp: return False
    pn, en = _norm(pred), _norm(exp)
    if not pn or not en: return False
    if pn == en or pn in en or en in pn: return True
    pw, ew = set(pn.split()), set(en.split())
    if len(pw) < 2 or len(ew) < 2: return False
    return len(pw & ew) / min(len(pw), len(ew)) >= 0.5

def compute_metrics(tp, fp, fn):
    p = tp / (tp + fp) if (tp + fp) else 0.0
    r = tp / (tp + fn) if (tp + fn) else 0.0
    f = 2*p*r/(p+r) if (p+r) else 0.0
    return {"precision": round(p, 4), "recall": round(r, 4), "f1": round(f, 4),
            "tp": tp, "fp": fp, "fn": fn}


# ═══════════════════════════════════════════════════════════════════
# 10. Main Pipeline
# ═══════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top-k", type=int, default=25)
    ap.add_argument("--threshold", type=float, default=0.15)
    ap.add_argument("--alpha", type=float, default=0.7, help="Semantic weight (1-alpha = BM25)")
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--skip-guardrail", action="store_true")
    ap.add_argument("--skip-sections", action="store_true")
    ap.add_argument("--skip-rerank", action="store_true", help="Skip cross-encoder reranking")
    ap.add_argument("--max-alerts", type=int, default=0)
    args = ap.parse_args()

    t0 = time.time()
    P = lambda *a, **kw: print(*a, **kw, flush=True)
    P("=" * 70)
    P("  POC-1 E2E TEST v3 — Multi-Query + Keyword + CrossEncoder")
    P("=" * 70)

    # ── 1. Ground truth ───────────────────────────────────────────
    P(f"\n[1/6] Parsing XLSX ground truth...")
    gt_entries, no_impact_lnis = parse_ground_truth(XLSX_PATH)
    alert_gt = build_alert_ground_truth(gt_entries)
    echo_to_gt_titles = {}
    for a in alert_gt.values():
        for eid, title in a["expected_docs"].items():
            echo_to_gt_titles[eid] = title
    total_gt_docs = sum(len(a["expected_docs"]) for a in alert_gt.values())
    total_gt_secs = sum(sum(len(v) for v in a["expected_sections"].values()) for a in alert_gt.values())
    P(f"  {len(alert_gt)} alerts | {total_gt_docs} doc mappings | {total_gt_secs} section mappings")

    # ── 2. Parse PG docs ──────────────────────────────────────────
    P(f"\n[2/6] Parsing PG docs from {PG_DOCS_DIR.name}...")
    xml_files = sorted(PG_DOCS_DIR.rglob("*.xml"))
    docs, echo_to_doc, errs = [], {}, 0
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futs = {pool.submit(parse_pg_xml, f): f for f in xml_files}
        for i, fut in enumerate(as_completed(futs)):
            r = fut.result()
            if r: docs.append(r); echo_to_doc[r.echo_id] = r
            else: errs += 1
            if (i+1) % 300 == 0: P(f"  {i+1}/{len(xml_files)}")
    gt_eids = set(echo_to_gt_titles); avail_eids = set(echo_to_doc)
    cov = gt_eids & avail_eids
    P(f"  {len(docs)} docs parsed ({errs} errors) | GT coverage: {len(cov)}/{len(gt_eids)}")

    # ── 3. Index (embeddings + BM25) ──────────────────────────────
    P(f"\n[3/6] Building indexes...")
    model = index_pg_docs(docs, TEST_QDRANT_DIR, batch_size=64)
    bm25_idx = BM25Index(docs)
    P(f"  BM25 index: {len(docs)} docs")
    t_index = time.time()
    P(f"  Indexing done in {t_index - t0:.1f}s")

    # ── 4. Retrieve + guardrail + sections ────────────────────────
    alerts_to_run = list(alert_gt.items())
    if args.max_alerts > 0: alerts_to_run = alerts_to_run[:args.max_alerts]
    P(f"\n[4/6] Running {len(alerts_to_run)} alerts "
      f"(top_k={args.top_k}, α={args.alpha}, thr={args.threshold})")

    results: list[AlertResult] = []
    total_groq = 0

    for i, (lni, gt) in enumerate(alerts_to_run):
        P(f"\n  [{i+1}/{len(alerts_to_run)}] {lni}")
        P(f"    {gt['title'][:65]}")
        P(f"    Terms: {gt['key_terms'][:65]}")
        P(f"    Expected: {len(gt['expected_docs'])} docs")

        hits = hybrid_retrieve(model, TEST_QDRANT_DIR, bm25_idx,
                               gt["title"], gt["key_terms"],
                               top_k=args.top_k * 2, threshold=args.threshold,
                               alpha=args.alpha, echo_to_doc=echo_to_doc)
        P(f"    Hybrid candidates: {len(hits)}")

        if not args.skip_rerank and hits:
            alert_text = f"{gt['title']} {gt['key_terms']}"
            hits = cross_encoder_rerank(alert_text, hits, echo_to_doc, top_k=args.top_k)
            P(f"    After cross-encoder rerank: {len(hits)}")

        if not args.skip_guardrail and hits:
            pre = len(hits)
            hits = guardrail_filter(gt["title"], gt["key_terms"], hits, echo_to_doc,
                                    workers=min(6, len(hits)))
            total_groq += pre
            P(f"    Guardrail: {pre} → {len(hits)} kept")

        ar = AlertResult(alert_lni=lni, alert_title=gt["title"], key_terms=gt["key_terms"],
                         expected_doc_ids=set(gt["expected_docs"].keys()),
                         expected_sections=gt["expected_sections"])
        ar.predicted_doc_ids = {h["echo_id"] for h in hits}
        ar.retrieval_scores = {h["echo_id"]: h["score"] for h in hits}
        ar.guardrail_kept = set(ar.predicted_doc_ids)
        ar.doc_tp = len(ar.predicted_doc_ids & ar.expected_doc_ids)
        ar.doc_fp = len(ar.predicted_doc_ids - ar.expected_doc_ids)
        ar.doc_fn = len(ar.expected_doc_ids - ar.predicted_doc_ids)
        P(f"    Doc TP={ar.doc_tp} FP={ar.doc_fp} FN={ar.doc_fn}")

        for h in hits[:5]:
            tag = "TP" if h["echo_id"] in ar.expected_doc_ids else "FP"
            P(f"      [{tag}] echo:{h['echo_id']:<10} "
              f"score={h['score']:.3f} (sem={h.get('sem_score',0):.3f} "
              f"bm25={h.get('bm25_score',0):.3f}) | {h['title'][:45]}")

        if not args.skip_sections:
            for h in hits:
                doc = echo_to_doc.get(h["echo_id"])
                if not doc or not doc.sections: continue
                matched = match_sections_sequential(gt["title"], gt["key_terms"],
                                                   doc.title, doc.sections)
                total_groq += min(len(doc.sections), 20)
                if matched:
                    ar.predicted_sections[h["echo_id"]] = [m["heading"] for m in matched]

            for eid, exp_heads in ar.expected_sections.items():
                pred_heads = ar.predicted_sections.get(eid, [])
                for eh in exp_heads:
                    if any(section_heading_matches(ph, eh) for ph in pred_heads):
                        ar.sec_tp += 1
                    else: ar.sec_fn += 1
            for eid, pred_heads in ar.predicted_sections.items():
                exp_heads = ar.expected_sections.get(eid, [])
                for ph in pred_heads:
                    if not any(section_heading_matches(ph, eh) for eh in exp_heads):
                        ar.sec_fp += 1
            P(f"    Sec TP={ar.sec_tp} FP={ar.sec_fp} FN={ar.sec_fn}")

        results.append(ar)

    t_ret = time.time()

    # ── 5. Aggregate ──────────────────────────────────────────────
    P(f"\n[5/6] Computing metrics & analysis...")
    d_tp = sum(r.doc_tp for r in results); d_fp = sum(r.doc_fp for r in results)
    d_fn = sum(r.doc_fn for r in results)
    dm = compute_metrics(d_tp, d_fp, d_fn)
    s_tp = sum(r.sec_tp for r in results); s_fp = sum(r.sec_fp for r in results)
    s_fn = sum(r.sec_fn for r in results)
    sm = compute_metrics(s_tp, s_fp, s_fn)

    # ── Score distribution analysis ──────────────────────────────
    all_tp_scores, all_fp_scores = [], []
    per_alert_analysis = []
    alerts_perfect_recall = 0
    alerts_zero_recall = 0
    alerts_with_fn_no_xml = 0

    for r in results:
        tp_scores = [r.retrieval_scores[e] for e in r.predicted_doc_ids & r.expected_doc_ids
                     if e in r.retrieval_scores]
        fp_scores = [r.retrieval_scores[e] for e in r.predicted_doc_ids - r.expected_doc_ids
                     if e in r.retrieval_scores]
        all_tp_scores.extend(tp_scores)
        all_fp_scores.extend(fp_scores)

        rm = compute_metrics(r.doc_tp, r.doc_fp, r.doc_fn)
        if rm["recall"] >= 1.0: alerts_perfect_recall += 1
        if rm["recall"] == 0.0 and (r.doc_tp + r.doc_fn) > 0: alerts_zero_recall += 1

        fn_missing_xml = sum(1 for e in r.expected_doc_ids - r.predicted_doc_ids
                             if e not in avail_eids)
        if fn_missing_xml > 0: alerts_with_fn_no_xml += 1

        score_gap = 0.0
        if tp_scores and fp_scores:
            score_gap = min(tp_scores) - max(fp_scores)

        per_alert_analysis.append({
            "alert_lni": r.alert_lni,
            "tp_score_min": round(min(tp_scores), 4) if tp_scores else None,
            "tp_score_max": round(max(tp_scores), 4) if tp_scores else None,
            "tp_score_mean": round(float(np.mean(tp_scores)), 4) if tp_scores else None,
            "fp_score_min": round(min(fp_scores), 4) if fp_scores else None,
            "fp_score_max": round(max(fp_scores), 4) if fp_scores else None,
            "fp_score_mean": round(float(np.mean(fp_scores)), 4) if fp_scores else None,
            "tp_fp_gap": round(score_gap, 4),
            "fn_missing_xml": fn_missing_xml,
        })

    score_analysis = {
        "tp_scores": {
            "count": len(all_tp_scores),
            "min": round(min(all_tp_scores), 4) if all_tp_scores else None,
            "max": round(max(all_tp_scores), 4) if all_tp_scores else None,
            "mean": round(float(np.mean(all_tp_scores)), 4) if all_tp_scores else None,
            "median": round(float(np.median(all_tp_scores)), 4) if all_tp_scores else None,
            "p25": round(float(np.percentile(all_tp_scores, 25)), 4) if all_tp_scores else None,
            "p75": round(float(np.percentile(all_tp_scores, 75)), 4) if all_tp_scores else None,
        },
        "fp_scores": {
            "count": len(all_fp_scores),
            "min": round(min(all_fp_scores), 4) if all_fp_scores else None,
            "max": round(max(all_fp_scores), 4) if all_fp_scores else None,
            "mean": round(float(np.mean(all_fp_scores)), 4) if all_fp_scores else None,
            "median": round(float(np.median(all_fp_scores)), 4) if all_fp_scores else None,
            "p25": round(float(np.percentile(all_fp_scores, 25)), 4) if all_fp_scores else None,
            "p75": round(float(np.percentile(all_fp_scores, 75)), 4) if all_fp_scores else None,
        },
        "overlap_zone": None,
    }
    if all_tp_scores and all_fp_scores:
        overlap_low = max(min(all_tp_scores), min(all_fp_scores))
        overlap_high = min(max(all_tp_scores), max(all_fp_scores))
        if overlap_low < overlap_high:
            score_analysis["overlap_zone"] = {
                "low": round(overlap_low, 4), "high": round(overlap_high, 4),
                "note": "TP and FP scores overlap in this range — hard to separate by threshold alone"
            }

    # ── Takeaways ─────────────────────────────────────────────────
    takeaways = []
    total_expected = d_tp + d_fn
    gt_missing = len(gt_eids - avail_eids)
    fn_from_missing = sum(a["fn_missing_xml"] for a in per_alert_analysis)

    takeaways.append(f"Ground truth has {total_expected} expected doc mappings across {len(results)} alerts.")
    takeaways.append(f"{gt_missing} of {len(gt_eids)} unique expected echo_ids have NO XML file "
                     f"({fn_from_missing} FN attributable to missing files).")
    takeaways.append(f"{alerts_perfect_recall}/{len(results)} alerts achieved 100% recall; "
                     f"{alerts_zero_recall}/{len(results)} alerts had 0% recall.")

    if all_tp_scores and all_fp_scores:
        takeaways.append(f"TP score range: {min(all_tp_scores):.3f}–{max(all_tp_scores):.3f} "
                         f"(mean {np.mean(all_tp_scores):.3f}). "
                         f"FP score range: {min(all_fp_scores):.3f}–{max(all_fp_scores):.3f} "
                         f"(mean {np.mean(all_fp_scores):.3f}).")
        if score_analysis["overlap_zone"]:
            takeaways.append(f"TP/FP overlap zone: {score_analysis['overlap_zone']['low']:.3f}–"
                             f"{score_analysis['overlap_zone']['high']:.3f}. "
                             f"Threshold tuning alone cannot fully separate TPs from FPs.")
        else:
            takeaways.append("No TP/FP score overlap — a threshold can cleanly separate TPs from FPs.")

    if d_fp > d_tp * 3:
        takeaways.append(f"PRECISION BOTTLENECK: {d_fp} FPs vs {d_tp} TPs. "
                         f"BM25 generic legal keywords (default, finance, security) likely cause noise.")
    if d_fn > total_expected * 0.4:
        takeaways.append(f"RECALL GAP: {d_fn}/{total_expected} expected docs missed. "
                         f"Embedding model may not capture domain-specific relevance for some topics.")

    # ── Build JSON report ─────────────────────────────────────────
    t_end = time.time()
    echo_info = {d.echo_id: {"doc_id": d.doc_id, "title": d.title} for d in docs}

    report = {
        "generated_at": datetime.now().isoformat(),
        "config": {"embedding": "all-MiniLM-L6-v2",
                   "cross_encoder": "ms-marco-MiniLM-L-6-v2" if not args.skip_rerank else "OFF",
                   "llm": GROQ_MODEL, "multi_query": True, "keyword_reranker": True,
                   "top_k": args.top_k, "threshold": args.threshold, "alpha": args.alpha,
                   "guardrail": not args.skip_guardrail, "sections": not args.skip_sections,
                   "pg_docs": len(docs), "alerts_tested": len(results)},
        "timing": {"total_s": round(t_end-t0,1), "index_s": round(t_index-t0,1),
                   "retrieval_s": round(t_ret-t_index,1), "groq_calls": total_groq},
        "doc_metrics": dm, "sec_metrics": sm,
        "score_analysis": score_analysis,
        "key_takeaways": takeaways,
        "gt_coverage": {"expected_unique_echo_ids": len(gt_eids),
                        "available_in_xml": len(cov),
                        "missing_count": gt_missing,
                        "missing_echo_ids": sorted(gt_eids - avail_eids),
                        "fn_from_missing_files": fn_from_missing},
        "alert_summary": {
            "total": len(results),
            "perfect_recall": alerts_perfect_recall,
            "zero_recall": alerts_zero_recall,
            "have_fn_due_to_missing_xml": alerts_with_fn_no_xml,
        },
        "per_alert": [],
    }
    for idx, r in enumerate(results):
        exp_rich = [{"echo_id": e, "doc_id": echo_info.get(e,{}).get("doc_id",""),
                     "title": echo_to_gt_titles.get(e,""),
                     "has_xml": e in avail_eids}
                    for e in sorted(r.expected_doc_ids)]
        pred_rich = [{"echo_id": e, "doc_id": echo_info.get(e,{}).get("doc_id",""),
                      "title": echo_info.get(e,{}).get("title","?"),
                      "score": r.retrieval_scores.get(e,0),
                      "correct": e in r.expected_doc_ids}
                     for e in sorted(r.retrieval_scores, key=lambda x: -r.retrieval_scores[x])]
        entry = {"alert_lni": r.alert_lni, "alert_title": r.alert_title,
                 "key_terms": r.key_terms, "expected_docs": exp_rich,
                 "predicted_docs": pred_rich,
                 "doc_tp": r.doc_tp, "doc_fp": r.doc_fp, "doc_fn": r.doc_fn,
                 "doc_metrics": compute_metrics(r.doc_tp, r.doc_fp, r.doc_fn),
                 "score_analysis": per_alert_analysis[idx]}
        if not args.skip_sections:
            entry["expected_sections"] = {k:v for k,v in r.expected_sections.items() if v}
            entry["predicted_sections"] = r.predicted_sections
            entry["sec_tp"] = r.sec_tp; entry["sec_fp"] = r.sec_fp; entry["sec_fn"] = r.sec_fn
            entry["sec_metrics"] = compute_metrics(r.sec_tp, r.sec_fp, r.sec_fn)
        report["per_alert"].append(entry)

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    rp = REPORT_DIR / f"test_report_v3_{ts}.json"
    with open(rp, "w") as f: json.dump(report, f, indent=2, ensure_ascii=False)

    # ── Print human-readable summary ──────────────────────────────
    P("\n" + "=" * 70)
    P("  E2E TEST REPORT v3")
    P("=" * 70)
    P(f"  Embedding : all-MiniLM-L6-v2 | CrossEncoder: {'ON' if not args.skip_rerank else 'OFF'}")
    P(f"  Pipeline  : Multi-Query + BM25(α={args.alpha}) + KeywordBoost + CrossEncoder")
    P(f"  top_k={args.top_k} | threshold={args.threshold}")
    P(f"  PG docs   : {len(docs)} | Alerts: {len(results)} | Groq calls: ~{total_groq}")
    P(f"  Time      : {t_end-t0:.1f}s (index={t_index-t0:.1f}s, retrieval={t_ret-t_index:.1f}s)")

    P(f"\n  DOCUMENT-LEVEL:")
    P(f"    Precision: {dm['precision']:.4f} ({d_tp}/{d_tp+d_fp})")
    P(f"    Recall   : {dm['recall']:.4f} ({d_tp}/{d_tp+d_fn})")
    P(f"    F1       : {dm['f1']:.4f}  |  TP={d_tp} FP={d_fp} FN={d_fn}")

    if not args.skip_sections:
        P(f"\n  SECTION-LEVEL:")
        P(f"    Precision: {sm['precision']:.4f} ({s_tp}/{s_tp+s_fp})")
        P(f"    Recall   : {sm['recall']:.4f} ({s_tp}/{s_tp+s_fn})")
        P(f"    F1       : {sm['f1']:.4f}  |  TP={s_tp} FP={s_fp} FN={s_fn}")

    P(f"\n  SCORE DISTRIBUTION (TP vs FP):")
    sa = score_analysis
    if sa["tp_scores"]["count"]:
        P(f"    TP scores : n={sa['tp_scores']['count']}  "
          f"min={sa['tp_scores']['min']}  median={sa['tp_scores']['median']}  "
          f"mean={sa['tp_scores']['mean']}  max={sa['tp_scores']['max']}  "
          f"[p25={sa['tp_scores']['p25']} p75={sa['tp_scores']['p75']}]")
    else:
        P(f"    TP scores : (none)")
    if sa["fp_scores"]["count"]:
        P(f"    FP scores : n={sa['fp_scores']['count']}  "
          f"min={sa['fp_scores']['min']}  median={sa['fp_scores']['median']}  "
          f"mean={sa['fp_scores']['mean']}  max={sa['fp_scores']['max']}  "
          f"[p25={sa['fp_scores']['p25']} p75={sa['fp_scores']['p75']}]")
    else:
        P(f"    FP scores : (none)")
    if sa["overlap_zone"]:
        P(f"    Overlap   : {sa['overlap_zone']['low']}–{sa['overlap_zone']['high']} "
          f"(cannot separate by threshold alone)")
    else:
        P(f"    Overlap   : NONE — clean threshold separation possible")

    P(f"\n  ALERT SUMMARY:")
    P(f"    Perfect recall (100%): {alerts_perfect_recall}/{len(results)}")
    P(f"    Zero recall (0%)    : {alerts_zero_recall}/{len(results)}")
    P(f"    FNs from missing XML: {fn_from_missing} across {alerts_with_fn_no_xml} alerts")

    P(f"\n  KEY TAKEAWAYS:")
    for j, t in enumerate(takeaways, 1):
        P(f"    {j}. {t}")

    P(f"\n  PER-ALERT (sorted by recall):")
    hdr = (f"  {'LNI':<36} {'Title':<28} {'P':>5} {'R':>5} {'F1':>5}"
           f" {'TP':>3} {'FP':>3} {'FN':>3}  {'TP_avg':>7} {'FP_avg':>7} {'Gap':>6}")
    P(hdr); P(f"  {'-'*(len(hdr)-2)}")
    for idx, r in enumerate(sorted(results,
            key=lambda x: compute_metrics(x.doc_tp,x.doc_fp,x.doc_fn)["recall"])):
        m = compute_metrics(r.doc_tp, r.doc_fp, r.doc_fn)
        pa = next(a for a in per_alert_analysis if a["alert_lni"] == r.alert_lni)
        tp_avg = f"{pa['tp_score_mean']:.3f}" if pa["tp_score_mean"] is not None else "   -  "
        fp_avg = f"{pa['fp_score_mean']:.3f}" if pa["fp_score_mean"] is not None else "   -  "
        gap = f"{pa['tp_fp_gap']:+.3f}" if pa["tp_score_mean"] is not None and pa["fp_score_mean"] is not None else "   -  "
        P(f"  {r.alert_lni:<36} {r.alert_title[:28]:<28} {m['precision']:>5.2f} "
          f"{m['recall']:>5.2f} {m['f1']:>5.2f} {r.doc_tp:>3} {r.doc_fp:>3} {r.doc_fn:>3}"
          f"  {tp_avg:>7} {fp_avg:>7} {gap:>6}")

    P(f"\n  MISSED DOCS (FN):")
    mc = 0
    for r in results:
        for eid in r.expected_doc_ids - r.predicted_doc_ids:
            mc += 1
            has_xml = "✓" if eid in avail_eids else "✗ NO XML"
            if mc <= 25:
                inf = echo_info.get(eid, {})
                P(f"    {r.alert_lni[:26]} | echo:{eid:<10} | "
                  f"doc:{inf.get('doc_id','N/A')[:30]} | {echo_to_gt_titles.get(eid,'?')[:35]} | {has_xml}")
    if mc > 25: P(f"    ... +{mc-25} more")

    P(f"\n  CORRECT RETRIEVALS (TP):")
    tc = 0
    for r in results:
        for eid in r.predicted_doc_ids & r.expected_doc_ids:
            tc += 1
            if tc <= 25:
                inf = echo_info.get(eid, {})
                P(f"    {r.alert_lni[:26]} | echo:{eid:<10} | "
                  f"doc:{inf.get('doc_id','?')[:30]} | {inf.get('title','?')[:35]} | "
                  f"score={r.retrieval_scores.get(eid,0):.3f}")
    if tc > 25: P(f"    ... +{tc-25} more")

    P(f"\n  Report: {rp}")
    P("=" * 70)

if __name__ == "__main__":
    main()
